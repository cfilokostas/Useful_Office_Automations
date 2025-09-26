import tkinter as tk
from tkinter import filedialog, messagebox
import time
import openpyxl
import os
from openpyxl import Workbook, load_workbook
import pandas as pd

def clock():
    date = time.strftime('%d/%m/%Y')
    currenttime = time.strftime('%H:%M:%S')
    datetimelabel.config(text=f'                Ημερομηνία: {date}\nΏρα:{currenttime}')
    datetimelabel.after(1000, clock)

def consolidate_excel_files(input_folder, output_file):
    # Create a new workbook to consolidate data
    consolidated_workbook = Workbook()

    # Iterate through each Excel file in the input folder
    for file in os.listdir(input_folder):
        if file.endswith(('.xlsx', '.xls')):
            file_path = os.path.join(input_folder, file)
            # Load the Excel file using openpyxl
            source_workbook = load_workbook(file_path)

            # Iterate through each sheet in the source workbook
            for sheet_name in source_workbook.sheetnames:
                source_sheet = source_workbook[sheet_name]
                # Create a new sheet in the consolidated workbook
                consolidated_sheet = consolidated_workbook.create_sheet(title=sheet_name)

                # Copy data from the source sheet into the consolidated sheet
                for row in source_sheet.iter_rows(values_only=True):
                    consolidated_sheet.append(row)

    # Save the consolidated workbook
    consolidated_workbook.save(output_file)


def ask_for_file():
    messagebox.showinfo("Info", "Παρακαλώ επιλέξτε το αρχείο Excel")
    file_path = filedialog.askopenfilename(title="Επιλογή αρχείου", filetypes=[("Excel Files", "*.xlsx *.xls")])
    return file_path

def ask_for_folder():
    messagebox.showinfo("Info", "Παρακαλώ επιλέξτε που να αποθηκευτούν τα αρχεία")
    folder_path = filedialog.askdirectory(title="Επιλογή φακέλου για αποθήκευση αρχείων")
    return folder_path

def ask_for_folder2():
    messagebox.showinfo("Info", "Παρακαλώ επιλέξτε φάκελο που περιέχει τα αρχεία")
    folder_path = filedialog.askdirectory(title="Επιλογή φακέλου για αποθήκευση αρχείων")
    return folder_path

def save_sheets_as_files(file_path, folder_path):
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheets = workbook.sheetnames

        for sheet_name in sheets:
            sheet = workbook[sheet_name]
            output_file = os.path.join(folder_path, f"{sheet_name}.xlsx")

            new_workbook = openpyxl.Workbook()
            new_sheet = new_workbook.active

            for row in sheet.iter_rows(values_only=True):
                new_sheet.append(row)

            new_workbook.save(output_file)

        messagebox.showinfo("Info", "Η διαδικασία ολοκληρώθηκε!")
    except Exception as e:
        print(f"An error occurred: {e}")

#####################################################################################
def koympi1():
    messagebox.showinfo("Info", "Κλέιστε την εφάρμογη 'Excel' εαν είναι ανοιχτή")
    input_file_path = ask_for_file()
    if not input_file_path:
        messagebox.showinfo("Info", "Δεν επιλέχθηκε αρχείο.")
        return

    output_folder_path = ask_for_folder()
    if not output_folder_path:
        messagebox.showinfo("Info", "Δεν επιλέχθηκε φάκελος.")
        return

    save_sheets_as_files(input_file_path, output_folder_path)
#####################################################################################
def koympi2():
    messagebox.showinfo("Info", "Κλέιστε την εφάρμογη 'Excel' εαν είναι ανοιχτή")
    input_folder_path = ask_for_folder2()
    output_file_name = 'enopoihmeno.xlsx'
    output_file_path = os.path.join(input_folder_path, output_file_name)
    consolidate_excel_files(input_folder_path, output_file_path)
    messagebox.showinfo("Info", "Η Διαδικασία Ολοκληρώθηκε")

#####################################################################################
def koympi3():
    messagebox.showinfo("Μαζική Δημιουργία Φακέλων", "Κλέιστε την εφάρμογη 'Excel' εαν είναι ανοιχτή")
    input_file_path = ask_for_file()
    if not input_file_path:
        messagebox.showinfo("Μαζική Δημιουργία Φακέλων", "Δεν επιλέχθηκε αρχείο.")
        return
    output_folder_path = ask_for_folder()

    workbook = openpyxl.load_workbook(input_file_path, read_only=True)
    sheet = workbook.active

    folder_counter = 0

    for row in sheet.iter_rows(values_only=True):
        value = row[0]
        if value is not None:
            folder_path = os.path.join(output_folder_path, str(value))
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)               
                folder_counter += 1
    messagebox.showinfo("Ολοκληρώθηκε", f"Δημιουργήθηκαν {folder_counter} φάκελοι.")



#####################################################################################
def koympi4():
    messagebox.showinfo("koympi4", "Not Ready yet")

def koympi5():
    messagebox.showinfo("Οδηγίες", "1) «Φύλλα σε αρχεία»\n\nΧωρίζει τα φύλλα από ένα αρχείο Excel και τα αποθηκεύει ως ξεχωριστά αρχεία.\n\nΒήματα:\n\ni) Επιλέγουμε ένα αρχείο excel\n\nii) Επιλέγουμε που τον φάκελο που θέλουμε να αποθηκευτούν τα φύλλα που δημιουργήθηκαν \n\n\n2) «Αρχεία σε Φύλλα»\n\nΑποθηκεύει τα βιβλία εργασίας ως φύλλα σε ένα ενοποιημένο αρχείο.\n\nΒήματα:\n\ni) Επιλέγουμε τον φάκελο που έχει τα αρχεία μας\n\n\n3) «Μαζική Δημιουργία Φακέλων»\n\n Δημιουργεί αυτόματα φακέλους και δίνει ως όνομα τις τιμές μιας λίστας από ένα αρχείο Excel.\n\nΒήματα:\n\ni)Δημιουργούμε ένα νέο βιβλίο εργασίας και βάζουμε στην πρώτη στήλη (Α) τις τιμές που θέλουμε να πάρουν ως όνομα οι φάκελοι\n\nii) Επιλέγουμε το αρχείο που δημιουργήσαμε.\n\niii)Επιλέγουμε που θα δημιουργηθούν οι φάκελοι.")

def koympi6():
    messagebox.showinfo("Πληροφορίες", "Version 1.0\nΑύγουστος 2023\n")

root = tk.Tk()
root.geometry('1280x700+0+0')
root.title('Εφαρμογή Διαχείρησης Αρχείων')
root.resizable(False, False)

datetimelabel = tk.Label(root, text='hello', font=('arial', 10, 'bold'))
datetimelabel.place(x=5, y=5)
clock()

s = 'Εφαρμογή Διαχείρησης Αρχείων '
sliderLabel = tk.Label(root, text=s, font=('arial', 18, 'bold'))
sliderLabel.place(x=340, y=0)



optionsframe = tk.Frame(root, bg='white')
optionsframe.place(x=260, y=120)

Button1 = tk.Button(optionsframe, text='Φύλλα σε αρχεία', width=25, bd=3, font=('arial', 18), bg='deepskyblue', fg='white',
                    command=koympi1)
Button1.grid(row=0, column=0, pady=0)

Button1 = tk.Button(optionsframe, text='Αρχεία σε Φύλλα', width=25, bd=3, font=('arial', 18), bg='deepskyblue', fg='white',
                    command=koympi2)
Button1.grid(row=0, column=1, pady=20)

Button1 = tk.Button(optionsframe, text='Μαζική Δημιουργία Φακέλων', width=25, bd=3, font=('arial', 18), bg='deepskyblue', fg='white',
                    command=koympi3)
Button1.grid(row=2, column=0, pady=2.5)

Button1 = tk.Button(optionsframe, text='-', width=25, bd=3, font=('arial', 18), bg='deepskyblue', fg='white',
                    command=koympi4)
Button1.grid(row=2, column=1, pady=20)

Button1 = tk.Button(optionsframe, text='Οδηγίες', width=25, bd=3, font=('arial', 18), bg='deepskyblue', fg='white',
                    command=koympi5)
Button1.grid(row=3, column=0, pady=2.5)

Button1 = tk.Button(optionsframe, text='Πληροφορίες', width=25, bd=3, font=('arial', 18), bg='deepskyblue', fg='white',
                    command=koympi6)
Button1.grid(row=3, column=1, pady=20)

root.mainloop()
